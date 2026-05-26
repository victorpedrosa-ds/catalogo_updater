"""
applier.py
----------
Aplica todas as mudanças aprovadas no catálogo Excel:
  - Preços         → insere novas linhas em PRECO-VIGENCIA
  - Novos produtos → insere em GTIN + PRODUTOS + PRECO-VIGENCIA
  - Removidos      → marca linhas em GTIN com cor e comentário
  - Descrições     → atualiza coluna MARCA/DESCRIÇÃO em PRODUTOS
Gera aba Relatório e salva como novo arquivo.
"""

import re
import shutil
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment

ABA_GTIN      = 'GTIN'
ABA_PRECO     = 'PRECO-VIGENCIA'
ABA_PRODUTOS  = 'PRODUTOS'
ABA_RELATORIO = 'Relatório'

COR_NOVO     = 'C6EFCE'   # verde claro
COR_REMOVIDO = 'FFC7CE'   # vermelho claro
COR_DESCRICAO = 'FFEB9C'  # amarelo claro


def aplicar_todas_mudancas(
    caminho_catalogo: str,
    mudancas_aprovadas:  list[dict],
    novos_aprovados:     list[dict],
    removidos_aprovados: list[dict],
    descricoes_aprovadas: list[dict],
) -> str:
    if not any([mudancas_aprovadas, novos_aprovados,
                removidos_aprovados, descricoes_aprovadas]):
        raise ValueError("Nenhuma mudança aprovada para aplicar.")

    caminho = Path(caminho_catalogo)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    caminho_saida = caminho.parent / f"{caminho.stem}_atualizado_{ts}{caminho.suffix}"
    shutil.copy2(str(caminho), str(caminho_saida))

    wb = load_workbook(str(caminho_saida))

    # 1. Inserir novos produtos (GTIN + PRODUTOS) e obter seus IDs
    ids_novos = {}
    if novos_aprovados:
        ids_novos = _inserir_novos_produtos(wb, novos_aprovados)

    # 2. Aplicar mudanças de preço (produtos existentes)
    if mudancas_aprovadas:
        _aplicar_precos(wb, mudancas_aprovadas)

    # 3. Inserir preço inicial dos novos produtos em PRECO-VIGENCIA
    if novos_aprovados and ids_novos:
        _inserir_precos_novos(wb, novos_aprovados, ids_novos)

    # 4. Marcar removidos na aba GTIN
    if removidos_aprovados:
        _marcar_removidos(wb, removidos_aprovados)

    # 5. Atualizar descrições na aba PRODUTOS
    if descricoes_aprovadas:
        _atualizar_descricoes(wb, descricoes_aprovadas)

    # 6. Gerar aba Relatório
    _gerar_relatorio(wb, mudancas_aprovadas, novos_aprovados,
                     removidos_aprovados, descricoes_aprovadas)

    wb.save(str(caminho_saida))
    print(f"[Applier] Arquivo salvo: {caminho_saida}")
    return str(caminho_saida)


# ════════════════════════════════════════════════════════════════════════════
# NOVOS PRODUTOS
# ════════════════════════════════════════════════════════════════════════════

def _inserir_novos_produtos(wb, novos_aprovados: list[dict]) -> dict:
    """
    Insere novos produtos nas abas GTIN e PRODUTOS. Retorna {gtin: id_produto}.
    Preenche todas as colunas com dados do PDF e replica formulas da linha anterior.
    Sem alteracao de formatacao.

    Estrutura confirmada do catalogo:
    GTIN  - Col1:ID_linha  Col2:GTIN  Col3:VALIDO  Col4:TIPO GTIN  Col5:Inclusao
             Col9:FATOR  Col11:ID PRODUTO  Cols12-19: formulas
    PRODUTOS - Col1:ID  Col2:N.GTINS(formula)  Col3:TIPO  Col4:FABRICANTE
               Col5:DESC CATALOGO(em branco)  Col6:DESC PORTARIA  Col7:EMBALAGEM
               Col8:VOLUME(ML)  Col9:MATERIAL  Col10:RETORNAVEL/DESCARTAVEL
               Col12:CONCATENAR(formula)  Col14:PRODUTO CALC(formula)
    """
    ws_gtin = wb[ABA_GTIN]
    ws_prod = wb[ABA_PRODUTOS] if ABA_PRODUTOS in wb.sheetnames else None

    # Colunas exatas da aba GTIN
    col_g_id_linha = _col_num(ws_gtin, ['ID'])
    col_g_gtin     = _col_num(ws_gtin, ['GTIN'])
    col_g_valido   = _col_num(ws_gtin, ['VALIDO', 'VALID'])
    col_g_tipo     = _col_num(ws_gtin, ['TIPO GTIN'])
    col_g_inclusao = _col_num(ws_gtin, ['Inclusao', 'INCLUSAO'])
    col_g_fator    = _col_num(ws_gtin, ['FATOR'])
    col_g_id_prod  = _col_num(ws_gtin, ['ID PRODUTO'])
    # Usa parcial para acentos nos headers do GTIN
    if col_g_valido is None:
        col_g_valido = _col_num_parcial(ws_gtin, ['VALID'])
    if col_g_inclusao is None:
        col_g_inclusao = _col_num_parcial(ws_gtin, ['INCLUS'])

    # Colunas exatas da aba PRODUTOS
    if ws_prod:
        col_p_id       = _col_num(ws_prod, ['ID'])
        col_p_fab      = _col_num(ws_prod, ['FABRICANTE'])
        col_p_portaria = _col_num_parcial(ws_prod, ['PORTARIA'])
        col_p_emb      = _col_num(ws_prod, ['EMBALAGEM'])
        col_p_vol      = _col_num_parcial(ws_prod, ['VOLUME'])
        col_p_mat      = _col_num(ws_prod, ['MATERIAL'])
        col_p_ret      = _col_num_parcial(ws_prod, ['RETORN', 'DESCART'])
    else:
        col_p_id = col_p_fab = col_p_portaria = None
        col_p_emb = col_p_vol = col_p_mat = col_p_ret = None

    proximo_id_produto = _proximo_id_disponivel(wb)
    ids_novos          = {}
    hoje               = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)

    for novo in novos_aprovados:
        gtin       = novo['gtin']
        nome       = novo.get('nome_pdf', '')
        fabricante = novo.get('fabricante', '')
        embalagem  = novo.get('embalagem', '')
        material   = novo.get('material', '')
        volume     = novo.get('volume', '')
        ret_desc      = novo.get('ret_desc', '')
        tipo_portaria = novo.get('tipo_portaria', 'REFRIGERANTE')
        id_produto    = proximo_id_produto

        ids_novos[gtin] = id_produto

        # Inserir na aba GTIN
        ultima_gtin = _ultima_linha_com_dado(ws_gtin)
        nova_g      = ultima_gtin + 1

        if col_g_id_linha:
            ultimo_id_g = ws_gtin.cell(ultima_gtin, col_g_id_linha).value
            try:
                proximo_id_g = int(str(ultimo_id_g).strip()) + 1
            except (TypeError, ValueError):
                proximo_id_g = nova_g
            ws_gtin.cell(nova_g, col_g_id_linha).value = proximo_id_g

        if col_g_gtin:
            ws_gtin.cell(nova_g, col_g_gtin).value    = gtin
        if col_g_valido:
            ws_gtin.cell(nova_g, col_g_valido).value  = True
        if col_g_tipo:
            # Usa o titulo exato do anexo para compor o TIPO GTIN.
            # Formato: 'GTIN PORTARIA - {TITULO DO ANEXO}'
            # Funciona para qualquer tipo presente ou futuro na portaria.
            if tipo_portaria:
                tipo_gtin_val = f'GTIN PORTARIA - {tipo_portaria}'
            else:
                # Fallback: copia da linha anterior
                tipo_anterior = ws_gtin.cell(ultima_gtin, col_g_tipo).value
                tipo_gtin_val = (tipo_anterior
                                 if tipo_anterior and not str(tipo_anterior).startswith('=')
                                 else None)
            if tipo_gtin_val:
                ws_gtin.cell(nova_g, col_g_tipo).value = tipo_gtin_val
        if col_g_inclusao:
            ws_gtin.cell(nova_g, col_g_inclusao).value = hoje
        if col_g_fator:
            ws_gtin.cell(nova_g, col_g_fator).value   = 1
        if col_g_id_prod:
            ws_gtin.cell(nova_g, col_g_id_prod).value = id_produto

        _replicar_formulas(ws_gtin, ultima_gtin, nova_g, colunas_preenchidas={
            c for c in [col_g_id_linha, col_g_gtin, col_g_valido,
                        col_g_tipo, col_g_inclusao, col_g_fator, col_g_id_prod]
            if c is not None
        })

        # Inserir na aba PRODUTOS
        if ws_prod:
            ultima_prod = _ultima_linha_com_dado(ws_prod)
            nova_p      = ultima_prod + 1

            if col_p_id:
                ws_prod.cell(nova_p, col_p_id).value       = id_produto
            # Col 3: TIPO — usa o titulo exato do anexo no PDF.
            # Aberto para qualquer tipo futuro sem necessidade de alterar o codigo.
            col_p_tipo = _col_num(ws_prod, ['TIPO'])
            if col_p_tipo and tipo_portaria:
                ws_prod.cell(nova_p, col_p_tipo).value = tipo_portaria

            if col_p_fab:
                ws_prod.cell(nova_p, col_p_fab).value      = fabricante or None
            if col_p_portaria:
                ws_prod.cell(nova_p, col_p_portaria).value = nome or None
            if col_p_emb:
                ws_prod.cell(nova_p, col_p_emb).value      = embalagem or None
            if col_p_vol and volume:
                try:
                    ws_prod.cell(nova_p, col_p_vol).value  = int(volume)
                except (ValueError, TypeError):
                    ws_prod.cell(nova_p, col_p_vol).value  = volume
            if col_p_mat:
                ws_prod.cell(nova_p, col_p_mat).value      = material or None
            if col_p_ret:
                ws_prod.cell(nova_p, col_p_ret).value      = ret_desc or None
            # Col 5 MARCA/DESCRICAO CATALOGO: em branco para preenchimento manual

            _replicar_formulas(ws_prod, ultima_prod, nova_p, colunas_preenchidas={
                c for c in [col_p_id, col_p_fab, col_p_portaria,
                             col_p_emb, col_p_vol, col_p_mat, col_p_ret]
                if c is not None
            })

        print(f"[Applier] Novo: GTIN {gtin} -> ID_PRODUTO {id_produto} | {nome}")
        proximo_id_produto += 1

    return ids_novos

def _inserir_precos_novos(wb, novos_aprovados: list[dict], ids_novos: dict):
    """Insere o preço inicial dos novos produtos em PRECO-VIGENCIA."""
    ws = wb[ABA_PRECO]

    ultima_linha = _ultima_linha_com_dado(ws)
    ultimo_id_row = ws.cell(ultima_linha, 1).value
    try:
        proximo_id_linha = int(ultimo_id_row) + 1
    except (TypeError, ValueError):
        proximo_id_linha = ultima_linha

    for novo in novos_aprovados:
        gtin = novo['gtin']
        if gtin not in ids_novos:
            continue

        id_produto = ids_novos[gtin]
        vigencia   = novo.get('vigencia', '')
        preco      = novo['preco']

        try:
            from datetime import datetime as dt
            vig_dt = dt.strptime(vigencia, '%d/%m/%Y')
        except Exception:
            vig_dt = vigencia

        nova_linha = ultima_linha + 1
        ultima_linha = nova_linha

        ws.cell(nova_linha, 1).value = proximo_id_linha
        ws.cell(nova_linha, 2).value = id_produto
        ws.cell(nova_linha, 3).value = vig_dt
        ws.cell(nova_linha, 4).value = None
        ws.cell(nova_linha, 5).value = preco

        for col in [6, 7, 8]:
            formula_ref = ws.cell(nova_linha - 1, col).value
            if formula_ref and str(formula_ref).startswith('='):
                nova_formula = _ajustar_formula(str(formula_ref), nova_linha - 1, nova_linha)
                ws.cell(nova_linha, col).value = nova_formula

        proximo_id_linha += 1


# ════════════════════════════════════════════════════════════════════════════
# MUDANÇAS DE PREÇO (produtos existentes)
# ════════════════════════════════════════════════════════════════════════════

def _aplicar_precos(wb, mudancas_aprovadas: list[dict]):
    ws = wb[ABA_PRECO]

    ultima_linha = _ultima_linha_com_dado(ws)
    ultimo_id_row = ws.cell(ultima_linha, 1).value
    try:
        proximo_id = int(ultimo_id_row) + 1
    except (TypeError, ValueError):
        proximo_id = ultima_linha

    for mudanca in mudancas_aprovadas:
        nova_linha   = ultima_linha + 1
        ultima_linha = nova_linha

        id_produto = mudanca['id_produto']
        vigencia   = mudanca['vigencia']
        preco_novo = mudanca['preco_novo']

        try:
            from datetime import datetime as dt
            vig_dt = dt.strptime(vigencia, '%d/%m/%Y')
        except Exception:
            vig_dt = vigencia

        ws.cell(nova_linha, 1).value = proximo_id
        ws.cell(nova_linha, 2).value = id_produto
        ws.cell(nova_linha, 3).value = vig_dt
        ws.cell(nova_linha, 4).value = None
        ws.cell(nova_linha, 5).value = preco_novo

        for col in [6, 7, 8]:
            formula_ref = ws.cell(nova_linha - 1, col).value
            if formula_ref and str(formula_ref).startswith('='):
                nova_formula = _ajustar_formula(str(formula_ref), nova_linha - 1, nova_linha)
                ws.cell(nova_linha, col).value = nova_formula

        proximo_id += 1
        print(f"[Applier] Preço: {mudanca['nome']} "
              f"R$ {mudanca['preco_atual']:.2f} → R$ {mudanca['preco_novo']:.2f}")


# ════════════════════════════════════════════════════════════════════════════
# REMOVIDOS
# ════════════════════════════════════════════════════════════════════════════

def _marcar_removidos(wb, removidos_aprovados: list[dict]):
    """Marca as linhas dos GTINs removidos na aba GTIN com cor e comentário."""
    ws_gtin = wb[ABA_GTIN]
    col_gtin_n = _col_num(ws_gtin, ['GTIN', 'GTIN / EAN', 'GTIN/EAN'])
    if not col_gtin_n:
        return

    gtins_remover = {re.sub(r'\D', '', str(r['gtin'])) for r in removidos_aprovados}
    fill_removido = PatternFill('solid', fgColor=COR_REMOVIDO)
    ts = datetime.now().strftime('%d/%m/%Y')

    for row in ws_gtin.iter_rows(min_row=2):
        cell_gtin = row[col_gtin_n - 1]
        gtin = re.sub(r'\D', '', str(cell_gtin.value or ''))
        if gtin in gtins_remover:
            _colorir_linha(ws_gtin, cell_gtin.row, fill_removido)
            try:
                cell_gtin.comment = Comment(
                    f'Removido da portaria em {ts}', 'PMPF Updater'
                )
            except Exception:
                pass
            print(f"[Applier] Removido marcado: GTIN {gtin}")


# ════════════════════════════════════════════════════════════════════════════
# ATUALIZAÇÃO DE DESCRIÇÃO
# ════════════════════════════════════════════════════════════════════════════

def _atualizar_descricoes(wb, descricoes_aprovadas: list[dict]):
    """
    Atualiza descrições na aba PRODUTOS:
      - MARCA/DESCRIÇÃO PORTARIA → nome exato da portaria (nome_novo)
      - MARCA/DESCRIÇÃO CATÁLOGO → nome digitado pelo usuário na interface (nome_catalogo)
                                    se em branco, mantém o valor anterior
    """
    if ABA_PRODUTOS not in wb.sheetnames:
        return

    ws_prod = wb[ABA_PRODUTOS]
    col_prod_id       = _col_num(ws_prod, ['ID'])
    col_prod_portaria = _col_num_parcial(ws_prod, ['PORTARIA'])
    col_prod_catalogo = _col_num_parcial(ws_prod, ['CATALOGO', 'CATÁLOGO'])

    if not col_prod_id or not col_prod_portaria:
        print("[Applier] Atenção: coluna ID ou MARCA/DESCRIÇÃO PORTARIA não encontrada em PRODUTOS.")
        return

    # {id_produto: {'nome_novo': ..., 'nome_catalogo': ...}}
    atualizacoes = {
        d['id_produto']: {
            'nome_novo':      d.get('nome_novo', ''),
            'nome_catalogo':  d.get('nome_catalogo', '').strip(),
        }
        for d in descricoes_aprovadas
    }

    for row in ws_prod.iter_rows(min_row=2):
        id_cell = row[col_prod_id - 1]
        try:
            id_prod = int(str(id_cell.value or '').strip())
        except (ValueError, TypeError):
            continue

        if id_prod not in atualizacoes:
            continue

        upd = atualizacoes[id_prod]

        # Atualiza coluna PORTARIA
        portaria_cell       = row[col_prod_portaria - 1]
        nome_antigo_port    = portaria_cell.value
        portaria_cell.value = upd['nome_novo']

        # Atualiza coluna CATÁLOGO (somente se o usuário digitou algo)
        nome_cat = upd['nome_catalogo']
        if col_prod_catalogo and nome_cat:
            catalogo_cell       = row[col_prod_catalogo - 1]
            nome_antigo_cat     = catalogo_cell.value
            catalogo_cell.value = nome_cat
            print(f"[Applier] ID {id_prod} — Portaria: '{nome_antigo_port}' → '{upd['nome_novo']}' | "
                  f"Catálogo: '{nome_antigo_cat}' → '{nome_cat}'")
        else:
            print(f"[Applier] ID {id_prod} — Portaria atualizada: '{nome_antigo_port}' → '{upd['nome_novo']}' "
                  f"(catálogo mantido)")



# ════════════════════════════════════════════════════════════════════════════
# RELATÓRIO
# ════════════════════════════════════════════════════════════════════════════

def _gerar_relatorio(wb, mudancas, novos, removidos, descricoes):
    if ABA_RELATORIO in wb.sheetnames:
        del wb[ABA_RELATORIO]
    ws = wb.create_sheet(ABA_RELATORIO)

    ts = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
    linha = 1

    def _titulo(texto, cor='2F4F4F'):
        ws.merge_cells(f'A{linha}:H{linha}')
        cel = ws[f'A{linha}']
        cel.value = texto
        cel.font  = Font(bold=True, size=11, color='FFFFFF')
        cel.fill  = PatternFill('solid', fgColor=cor)
        cel.alignment = Alignment(horizontal='center')

    def _cabecalho(cols):
        for ci, cab in enumerate(cols, 1):
            cel = ws.cell(row=linha, column=ci, value=cab)
            cel.font = Font(bold=True, color='FFFFFF')
            cel.fill = PatternFill('solid', fgColor='555555')
            cel.alignment = Alignment(horizontal='center')

    # Cabeçalho geral
    ws.merge_cells(f'A{linha}:H{linha}')
    cel = ws[f'A{linha}']
    cel.value = f'Relatório PMPF Updater — {ts}'
    cel.font  = Font(bold=True, size=13)
    cel.alignment = Alignment(horizontal='center')
    linha += 2

    # ── Mudanças de preço ──────────────────────────────────────────────────
    if mudancas:
        _titulo(f'MUDANÇAS DE PREÇO ({len(mudancas)} itens)', '2F4F4F')
        linha += 1
        _cabecalho(['GTIN', 'ID', 'PRODUTO', 'VIGÊNCIA', 'PREÇO ANTERIOR', 'PREÇO NOVO', 'VARIAÇÃO', ''])
        linha += 1
        for m in mudancas:
            variacao = m['preco_novo'] - m['preco_atual']
            cor = 'FFC7CE' if variacao > 0 else 'C6EFCE'
            vals = [m['gtin'], m['id_produto'], m['nome'], m['vigencia'],
                    round(m['preco_atual'], 2), round(m['preco_novo'], 2), round(variacao, 2), '']
            for ci, v in enumerate(vals, 1):
                cel = ws.cell(row=linha, column=ci, value=v)
                if ci >= 5:
                    cel.fill = PatternFill('solid', fgColor=cor)
            linha += 1
        linha += 1

    # ── Novos produtos ─────────────────────────────────────────────────────
    if novos:
        _titulo(f'NOVOS PRODUTOS ({len(novos)} itens)', '1F6E3C')
        linha += 1
        _cabecalho(['GTIN', 'NOME (PDF)', 'FABRICANTE', 'EMBALAGEM', 'VOLUME', 'PREÇO', 'VIGÊNCIA', ''])
        linha += 1
        for p in novos:
            vals = [p['gtin'], p.get('nome_pdf',''), p.get('fabricante',''),
                    p.get('embalagem',''), p.get('volume',''),
                    round(p['preco'], 2), p.get('vigencia',''), '']
            for ci, v in enumerate(vals, 1):
                cel = ws.cell(row=linha, column=ci, value=v)
                cel.fill = PatternFill('solid', fgColor=COR_NOVO)
            linha += 1
        linha += 1

    # ── Removidos ──────────────────────────────────────────────────────────
    if removidos:
        _titulo(f'REMOVIDOS DA PORTARIA ({len(removidos)} itens)', '8B0000')
        linha += 1
        _cabecalho(['GTIN', 'ID', 'PRODUTO', 'ÚLTIMO PREÇO', 'ÚLTIMA VIGÊNCIA', '', '', ''])
        linha += 1
        for r in removidos:
            preco_str = round(r['ultimo_preco'], 2) if r['ultimo_preco'] else ''
            vals = [r['gtin'], r['id_produto'], r['nome'],
                    preco_str, r.get('ultima_vigencia',''), '', '', '']
            for ci, v in enumerate(vals, 1):
                cel = ws.cell(row=linha, column=ci, value=v)
                cel.fill = PatternFill('solid', fgColor=COR_REMOVIDO)
            linha += 1
        linha += 1

    # ── Descrições ─────────────────────────────────────────────────────────
    if descricoes:
        _titulo(f'ATUALIZAÇÕES DE DESCRIÇÃO ({len(descricoes)} itens)', '7B4F00')
        linha += 1
        _cabecalho(['GTIN', 'ID', 'DESCRIÇÃO ANTERIOR', 'NOVA DESCRIÇÃO (PORTARIA)',
                    'SIMILARIDADE', '', '', ''])
        linha += 1
        for d in descricoes:
            vals = [d['gtin'], d['id_produto'], d.get('nome_atual',''),
                    d.get('nome_novo',''), f"{d.get('similaridade',0):.0f}%", '', '', '']
            for ci, v in enumerate(vals, 1):
                cel = ws.cell(row=linha, column=ci, value=v)
                cel.fill = PatternFill('solid', fgColor=COR_DESCRICAO)
            linha += 1

    # Larguras
    for ci, larg in enumerate([20, 10, 45, 14, 18, 18, 14, 5], 1):
        ws.column_dimensions[get_column_letter(ci)].width = larg

    # Resumo
    linha += 1
    resumo = (f"Total aplicado: {len(mudancas)} preços | "
              f"{len(novos)} novos | {len(removidos)} removidos | "
              f"{len(descricoes)} descrições")
    ws.cell(linha, 1, resumo).font = Font(bold=True)


# ════════════════════════════════════════════════════════════════════════════
# UTILITÁRIOS
# ════════════════════════════════════════════════════════════════════════════

def _col_num(ws, candidatos: list) -> int | None:
    """Retorna o número (1-indexed) da coluna cujo cabeçalho bate com um dos candidatos.
    Normaliza Unicode (NFC) antes de comparar para evitar falsos negativos com acentos."""
    import unicodedata
    if ws is None:
        return None
    for row in ws.iter_rows(min_row=1, max_row=3):
        for cell in row:
            if cell.value is not None:
                val = unicodedata.normalize('NFC', str(cell.value)).strip().upper()
                for cand in candidatos:
                    if val == unicodedata.normalize('NFC', cand).strip().upper():
                        return cell.column
    return None


def _ultima_linha_com_dado(ws) -> int:
    ultima = ws.max_row
    while ultima > 1 and ws.cell(ultima, 1).value is None:
        ultima -= 1
    return ultima


def _proximo_id_disponivel(wb) -> int:
    max_id = 0
    for aba, candidatos in [
        (ABA_GTIN,     ['ID PRODUTO']),
        (ABA_PRODUTOS, ['ID']),
    ]:
        if aba not in wb.sheetnames:
            continue
        ws  = wb[aba]
        col = _col_num(ws, candidatos)
        if not col:
            continue
        for row in ws.iter_rows(min_row=2):
            val = row[col - 1].value
            try:
                max_id = max(max_id, int(str(val).strip()))
            except (TypeError, ValueError, AttributeError):
                pass
    return max_id + 1


def _colorir_linha(ws, num_linha: int, fill: PatternFill):
    for col in range(1, ws.max_column + 1):
        ws.cell(num_linha, col).fill = fill


def _ajustar_formula(formula: str, linha_origem: int, linha_destino: int) -> str:
    import re
    def substituir(m):
        col_letra = m.group(1)
        num_linha = int(m.group(2))
        if num_linha == linha_origem:
            return f"{col_letra}{linha_destino}"
        return m.group(0)
    return re.sub(r'([A-Z]+)(\d+)', substituir, formula)

def _replicar_formulas(ws, linha_origem: int, linha_destino: int, colunas_preenchidas: set):
    """
    Para cada coluna da linha_origem que contenha uma fórmula (começa com '=')
    e cujo número NÃO esteja em colunas_preenchidas, copia a fórmula ajustada
    para a linha_destino. Isso garante que fórmulas como CONT.SE (coluna B)
    sejam replicadas automaticamente em novas linhas.
    """
    for col in range(1, ws.max_column + 1):
        if col in colunas_preenchidas:
            continue
        cell_origem = ws.cell(linha_origem, col)
        valor = cell_origem.value
        if valor and str(valor).startswith('='):
            nova_formula = _ajustar_formula(str(valor), linha_origem, linha_destino)
            ws.cell(linha_destino, col).value = nova_formula


def _col_num_parcial(ws, fragmentos: list) -> int | None:
    """
    Retorna o número (1-indexed) da primeira coluna cujo cabeçalho contenha
    pelo menos um dos fragmentos (case-insensitive, normalizado Unicode).
    """
    import unicodedata
    if ws is None:
        return None
    for row in ws.iter_rows(min_row=1, max_row=3):
        for cell in row:
            if cell.value is not None:
                val = unicodedata.normalize('NFC', str(cell.value)).strip().upper()
                for frag in fragmentos:
                    if unicodedata.normalize('NFC', frag).upper() in val:
                        return cell.column
    return None
